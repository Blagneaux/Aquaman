import os
import serial
import serial.tools.list_ports
import threading
import tkinter as tk
from tkinter import filedialog, messagebox
from openpyxl import Workbook, load_workbook
import matplotlib.pyplot as plt
from datetime import datetime
from openpyxl.styles import NamedStyle
import pandas as pd
from tkinter import ttk


class SerialReaderApp:
    def __init__(self, master):
        self.master = master
        master.title("Lecteur Série")

        self.serial_port = None
        self.is_reading = False
        self.data_filename = None
        self.data_buffer = []

        self.create_widgets()

    def create_widgets(self):
        style = ttk.Style()
        style.configure("TButton", padding=6, relief="flat")

        self.frame = ttk.Frame(self.master, padding=(20, 10))
        self.frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))

        self.start_button = ttk.Button(
            self.frame, text="Démarrer Lecture", command=self.start_reading
        )
        self.start_button.grid(row=0, column=0, pady=10, padx=10, sticky=tk.W)

        self.stop_button = ttk.Button(
            self.frame,
            text="Arrêter Lecture",
            command=self.stop_reading,
            state=tk.DISABLED,
        )
        self.stop_button.grid(row=0, column=1, pady=10, padx=10, sticky=tk.W)

    def start_reading(self):
        if not self.is_reading:
            file_path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Fichiers Excel", "*.xlsx"), ("Tous les fichiers", "*.*")],
            )
            if file_path:
                self.data_filename = file_path
                self.serial_port = self.recup_port_Arduino()
                if self.serial_port:
                    self.is_reading = True
                    self.set_button_state(self.start_button, tk.DISABLED)
                    self.set_button_state(self.stop_button, tk.NORMAL)
                    threading.Thread(target=self.read_serial).start()
                else:
                    messagebox.showerror("Erreur", "Port Arduino non trouvé.")
            else:
                messagebox.showinfo(
                    "Info", "Sélection d'emplacement de sauvegarde annulée."
                )
        else:
            messagebox.showinfo("Info", "La lecture est déjà en cours.")

    def stop_reading(self):
        if self.is_reading and messagebox.askyesno(
            "Confirmation", "Voulez-vous vraiment arrêter la lecture ?"
        ):
            self.is_reading = False
            self.serial_port.close()
            self.set_button_state(self.start_button, tk.NORMAL)
            self.set_button_state(self.stop_button, tk.DISABLED)
            self.write_to_excel()
            self.show_plot()
            messagebox.showinfo("Succès", "Lecture arrêtée avec succès.")
        else:
            messagebox.showinfo("Info", "La lecture n'est pas en cours.")

    def read_serial(self):
        while self.is_reading:
            try:
                if self.serial_port and self.serial_port.is_open:
                    data = self.serial_port.readline().decode("utf-8", errors="replace")
                    print("data: ",data.strip())
                    pressure_data, timestamp = self.extract_data(data)
                    if pressure_data is not None and timestamp is not None:
                        self.data_buffer.append((pressure_data, timestamp))
            except (serial.SerialException, UnicodeDecodeError) as e:
                print(f"Erreur de lecture depuis le port série : {e}")
                self.stop_reading()

    def extract_data(self, data):
        parts = data.split()
        if len(parts) == 9:
            try:
                pressure_data = [float(part) for part in parts[:8]]
                timestamp = float(parts[-1])
                return pressure_data, timestamp
            except ValueError:
                print(f"Erreur de conversion des données : {data}")
                return None, None
        else:
            return None, None

    def show_plot(self):
        if self.data_filename and os.path.exists(self.data_filename):
            # Charger les données depuis le fichier Excel
            df = pd.read_excel(self.data_filename)

            # Afficher les noms de colonnes
            print("Noms de colonnes dans le fichier Excel :")
            print(df.columns)

            # Créer un graphique avec les 8 courbes de chaque capteur
            plt.figure(figsize=(10, 6))

            # Utiliser l'ordre correct des capteurs
            capteur_order = [1, 2, 3, 7, 4, 5, 6, 8]
            new_legend_names = ['Capteur 1', 'Capteur 2', 'Capteur 3', 'Capteur 4', 'Capteur 5', 'Capteur 6', 'Capteur 7', 'Capteur 8']


            # Utiliser les indices de colonnes pour les capteurs (colonnes 1 à 8)
            for i in capteur_order:
                pressure_values = df.iloc[:, i - 1]
                converted_values = self.convert_pressure_values(pressure_values)
                plt.plot(df.index, converted_values, label=f"Capteur {i}")

            # Ajouter des étiquettes et une légende
            plt.xlabel("Timer")
            plt.ylabel("Pression (Pascal)")
            plt.title("Courbes de pression pour chaque capteur")
            plt.legend(new_legend_names)

            # Afficher le graphique
            plt.show()
        else:
            messagebox.showwarning("Attention", "Aucun fichier de données trouvé.")

    def convert_pressure_values(self, pressure_values):
        # Adapter les valeurs de pression entre -500 et 500 Pascal
        min_binary_value = -32768
        max_binary_value = 32768
        min_pressure_value = -500
        max_pressure_value = 500

        converted_values = (pressure_values - min_binary_value) * (
            max_pressure_value - min_pressure_value
        ) / (max_binary_value - min_binary_value) + min_pressure_value

        return converted_values

    def save_file(self):
        if self.data_buffer:
            file_path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Fichiers Excel", "*.xlsx"), ("Tous les fichiers", "*.*")],
            )
            if file_path:
                self.data_filename = file_path
                self.write_to_excel()
                messagebox.showinfo(
                    "Succès", f"Le fichier a été sauvegardé sous {self.data_filename}"
                )
            else:
                messagebox.showinfo(
                    "Info", "Sélection d'emplacement de sauvegarde annulée."
                )
        else:
            messagebox.showinfo("Info", "Aucune donnée à sauvegarder.")

    def write_to_excel(self):
        if self.data_filename:
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Serial Data"

            integer_style = NamedStyle(name="integer", number_format="0")

            for row_index, (pressure_data, timestamp) in enumerate(
                self.data_buffer, start=1
            ):
                for col_index, value in enumerate(pressure_data + [timestamp], start=1):
                    cleaned_value = int(value)
                    sheet.cell(
                        row=row_index, column=col_index, value=cleaned_value
                    ).style = integer_style

            workbook.save(self.data_filename)
            self.data_buffer = []

    @staticmethod
    def set_button_state(button, state):
        button.config(state=state)

    @staticmethod
    def recup_port_Arduino():
        ports = list(serial.tools.list_ports.comports())
        for p in ports:
            if "USB-SERIAL CH340" in p.description:
                return serial.Serial(p.device, 1000000)
        return None


if __name__ == "__main__":
    root = tk.Tk()
    app = SerialReaderApp(root)
    root.mainloop()
